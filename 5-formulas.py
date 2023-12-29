from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Lê uma pasta de trabalho
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatório"]

# Referências das linhas e colunas
min_column = sheet.min_column
max_column = sheet.max_column
min_row = sheet.min_row
max_row = sheet.max_row

# # - 3 incluindo formulas
# sheet["B6"] = "=SUM(B2:B5)"
# sheet["B6"].style = "Currency"
for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"


wb.save("test.xlsx")